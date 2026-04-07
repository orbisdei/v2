import openpyxl, json, re, sys

wb = openpyxl.load_workbook(r'C:\Users\User\Documents\orbis-dei\Orbis Dei Database v3.xlsm', read_only=True, data_only=True, keep_vba=False)

# ── Build lookups ─────────────────────────────────────────────────────────────

country_lookup = {}
for row in wb['Country Tags'].iter_rows(min_row=6, values_only=True):
    if row[3] and row[6]:
        country_lookup[str(row[3]).strip()] = str(row[6]).strip()

local_lookup = {}
for row in wb['Local Tags'].iter_rows(min_row=6, values_only=True):
    if row[0] == 'Done' and row[2]:
        slug = str(row[2]).strip()
        local_lookup[slug] = {
            'id': slug,
            'name': str(row[3]).strip() if row[3] else slug,
            'description': str(row[5]).strip() if row[5] else '',
            'featured': False
        }

cat_lookup = {}
for row in wb['Tags'].iter_rows(min_row=6, values_only=True):
    if row[2]:
        slug = str(row[2]).strip()
        name = str(row[3]).strip() if row[3] and row[3] != 'Page Title (what will display on website)' else slug
        cat_lookup[slug] = {
            'id': slug,
            'name': name,
            'description': str(row[5]).strip() if row[5] else '',
            'featured': bool(row[18]) if row[18] is not None else False
        }

ws_places = wb['Places']
hdr = list(ws_places.iter_rows(min_row=5, max_row=5, values_only=True))[0]
AL_NAME = str(hdr[37]).strip() if hdr[37] else 'The Real Presence'
AN_NAME = str(hdr[39]).strip() if hdr[39] else 'Miracle Hunter'
AP_NAME = str(hdr[41]).strip() if hdr[41] else 'MaryPages'

NON_CITY_TAGS = set(cat_lookup.keys()) | {'holy-persons'}

# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(text):
    t = str(text).lower().strip()
    for a, b in [("'", ''), ('\u2019', ''), ('\u2018', ''), ('`', ''),
                 ('e\u0301', 'e'), ('\xe9', 'e'), ('\xe8', 'e'), ('\xea', 'e'), ('\xeb', 'e'),
                 ('\xe0', 'a'), ('\xe2', 'a'), ('\xe4', 'a'),
                 ('\xf1', 'n'), ('\xf3', 'o'), ('\xf6', 'o'), ('\xfa', 'u'), ('\xfc', 'u'),
                 ('\u0159', 'r'), ('\u0161', 's'), ('\u017e', 'z'), ('\u010d', 'c'),
                 ('\xfd', 'y'), ('\xed', 'i'), ('\xe1', 'a'), ('\u010d', 'c')]:
        t = t.replace(a, b)
    t = re.sub(r'[^a-z0-9\s-]', '', t)
    t = re.sub(r'[\s-]+', '-', t)
    return t.strip('-')

def parse_tags(bl_value):
    if not bl_value:
        return []
    tags = []
    for line in str(bl_value).split('\n'):
        line = re.sub(r'_x000D_', '', line).strip()
        if line.startswith('- '):
            tag = line[2:].strip()
            if tag:
                tags.append(tag)
    return tags

def clean_desc(text):
    if not text:
        return ''
    text = re.sub(r'!\[.*?\]\(.*?\)', '', text)
    return text.strip()

def v(cell):
    if cell is None or cell == "'":
        return None
    s = str(cell).strip()
    return s if s not in ('', "'") else None

def maps_url(title, city, country):
    name = re.sub(r'\s*\(.*?\)', '', title).strip()
    parts = [p for p in [name, city, country] if p]
    q = '+'.join(p.replace(' ', '+') for p in parts)
    return 'https://maps.google.com/?q=' + q

# ── Process rows ──────────────────────────────────────────────────────────────

sites_out = []
all_tag_ids_used = set()

for row_num in range(8, 300):
    rows = list(ws_places.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    if not rows:
        break
    row = rows[0]
    if row[0] is None:
        break
    if row[0] != 'In Progress':
        continue

    contributor  = v(row[1])
    country_name = v(row[2]) or ''
    local_tag_d  = v(row[3])
    lat          = row[4]
    lon          = row[5]
    full_title   = v(row[6]) or ''
    interest     = v(row[8])
    description  = clean_desc(v(row[12]))
    author_note  = v(row[13])
    official_url = v(row[14])
    official_cmt = v(row[15])
    l1n = v(row[16]); l1u = v(row[17]); l1c = v(row[18])
    l2n = v(row[19]); l2u = v(row[20]); l2c = v(row[21])
    l3n = v(row[22]); l3u = v(row[23]); l3c = v(row[24])
    year    = v(row[36])
    al_url  = v(row[37]); al_cmt = v(row[38])
    an_url  = v(row[39]); an_cmt = v(row[40])
    ap_url  = v(row[41]); ap_cmt = v(row[42])
    bl_raw  = row[63]

    tags = parse_tags(bl_raw)
    country_code = tags[0] if tags and len(tags[0]) == 2 else ''
    all_tag_ids_used.update(tags)

    if local_tag_d:
        city_slug = slugify(local_tag_d)
        city_display = local_tag_d
    else:
        city_slug = None
        city_display = None
        # Only treat a BL tag as city if it's explicitly in the Local Tags sheet
        for t in tags[1:]:
            if t in local_lookup:
                city_slug = t
                city_display = local_lookup[t]['name']
                break

    title_slug = slugify(full_title)
    if city_slug:
        site_id = f"{country_code}-{city_slug}-{title_slug}"
    else:
        site_id = f"{country_code}-{title_slug}"
    if len(site_id) > 80:
        site_id = site_id[:80].rstrip('-')

    links = []
    if official_url:
        lnk = {'url': official_url, 'link_type': 'Official Website'}
        if official_cmt:
            lnk['comment'] = official_cmt
        links.append(lnk)
    for (n, u, c) in [(l1n, l1u, l1c), (l2n, l2u, l2c), (l3n, l3u, l3c)]:
        if n and u:
            lnk = {'url': u, 'link_type': n}
            if c:
                lnk['comment'] = c
            links.append(lnk)
    for (name, u, c) in [(AL_NAME, al_url, al_cmt), (AN_NAME, an_url, an_cmt), (AP_NAME, ap_url, ap_cmt)]:
        if u:
            lnk = {'url': u, 'link_type': name}
            if c:
                lnk['comment'] = c.strip()
            links.append(lnk)

    site = {
        'id': site_id,
        'name': full_title,
        'short_description': description,
        'latitude': lat,
        'longitude': lon,
        'google_maps_url': maps_url(full_title, city_display or '', country_name),
        'featured': False,
        'interest': interest,
        'contributor': contributor,
        'contributor_notes': [author_note] if author_note else [],
        'updated_at': '2026-03-16T00:00:00Z',
        'images': [],
        'links': links,
        'tag_ids': tags
    }
    if year:
        site['founded_date'] = str(year)

    sites_out.append(site)

# ── Build new tags ────────────────────────────────────────────────────────────

existing_tag_ids = {'therese-lisieux', 'churches', 'marian-sites', 'ar', 'be', 'brussels'}
tags_out = []

for tid in sorted(all_tag_ids_used):
    if tid in existing_tag_ids:
        continue
    if len(tid) == 2 and tid in country_lookup:
        tags_out.append({'id': tid, 'name': country_lookup[tid],
                         'description': f'Holy sites in {country_lookup[tid]}.', 'featured': False})
    elif tid in local_lookup:
        tags_out.append(local_lookup[tid])
    elif tid in cat_lookup:
        tags_out.append(cat_lookup[tid])
    else:
        tags_out.append({'id': tid, 'name': tid.replace('-', ' ').title(), 'description': '', 'featured': False})

with open(r'C:\Users\User\Documents\orbis-dei\scripts\import_output.json', 'w', encoding='utf-8') as f:
    json.dump({'new_sites': sites_out, 'new_tags': tags_out}, f, indent=2, ensure_ascii=False)

print(f"Done: {len(sites_out)} sites, {len(tags_out)} new tags")
